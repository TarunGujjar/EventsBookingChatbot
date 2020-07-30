import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('event_list.db')
wb = load_workbook('Events_details.xlsx')
ws = wb['Events_info']

conn.execute("create table if not exists events_details (Events_name text, Events_description text, Price int)")

for i in range(1, 7):
    temp_str = 'insert into events_details (Events_name, Events_description, Price) values ("{0}", "{1}", "{2}")'\
        .format(ws.cell(i, 2).value, ws.cell(i, 3).value, ws.cell(i, 4).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from events_details")
for i in content:
    print(i)
